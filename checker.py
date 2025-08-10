from openpyxl import load_workbook

class ExcelChecker:
    def __init__(self):
        pass

    def get_sheet_names(self, file_path):
        """获取Excel文件中的所有sheet名称"""
        try:
            wb = load_workbook(file_path, read_only=True)
            return wb.sheetnames
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return []

    def execute_check(self, option_id, file1, sheet1, file2, sheet2):
        """
        执行校验
        :param option_id: 校验选项ID
        :param file1: 文件1路径
        :param sheet1: 文件1的sheet名称
        :param file2: 文件2路径
        :param sheet2: 文件2的sheet名称
        :return: 校验结果列表 [(项目, 文件1值, 文件2值, 是否通过), ...]
        """
        results = []
        try:
            # 加载工作簿
            wb1 = load_workbook(file1, data_only=True)
            ws1 = wb1[sheet1]
            wb2 = load_workbook(file2, data_only=True)
            ws2 = wb2[sheet2]

            # 根据选项ID执行不同的校验
            if option_id == 1:
                results = self.check_loan_items(ws1, ws2)
            elif option_id == 2:
                results = self.check_overdue_loans(ws1, ws2)  # 逾期贷款核对(G01和S6301)
            elif option_id == 3:
                results = self.check_overdue_60_days(ws1, ws2)  # 逾期贷款(60天以上)核对
            elif option_id == 4:
                results = self.check_overdue_90_days(ws1, ws2)  # 逾期贷款(90天以上)核对
            elif option_id == 5:
                results = self.check_overdue_days(ws1, ws2)  # 各逾期天数贷款核对(S6301和S7101)
            elif option_id == 7:
                results = self.check_housing_loan_balance(ws1, ws2)
            elif option_id == 8:
                results = self.check_overdue_loan_1(ws1, ws2)
            elif option_id == 9:
                results = self.check_overdue_loan_2(ws1, ws2)
            elif option_id == 10:
                results = self.check_overdue_loan_3(ws1, ws2)
            elif option_id == 11:
                results = self.check_housing_loan_ratio_1(ws1, ws2)
            elif option_id == 12:
                results = self.check_housing_loan_ratio_2(ws1, ws2)
            elif option_id == 13:
                results = self.check_overdue_days(ws1, ws2)

            return results
        except Exception as e:
            return [('错误', str(e), '', False)]

    def execute_four_check(self, option_id, file1, sheet1, file2, sheet2, file3, sheet3, file4, sheet4):
        results = []
        try:
            # 加载四个工作簿
            wb1 = load_workbook(file1, data_only=True)
            ws1 = wb1[sheet1]
            wb2 = load_workbook(file2, data_only=True)
            ws2 = wb2[sheet2]
            wb3 = load_workbook(file3, data_only=True)
            ws3 = wb3[sheet3]
            wb4 = load_workbook(file4, data_only=True)
            ws4 = wb4[sheet4]

            # 根据选项ID执行不同的校验
            if option_id == 1:
                results = self.check_interest_rate_abnormal(ws1, ws2, ws3, ws4)
            elif option_id == 2:
                results = self.check_loan_provision_coverage(ws1, ws2, ws3, ws4)

            return results
        except Exception as e:
            return [('错误', str(e), '', '', False)]

    def check_interest_rate_abnormal(self, ws1, ws2, ws3, ws4):
        """利率异常校验：
        校验表1中C6的数据-表2中C6的数据的差除以表2中C6的数据大于20%，
        同时表3中C9的数据-表4中C9的数据差除以表4中C9的数据大于5%，
        此时校验结果为异常
        """
        results = []

        # 获取单元格值
        c6_ws1 = self.get_cell_value(ws1, 'C6')
        c6_ws2 = self.get_cell_value(ws2, 'C6')
        c9_ws3 = self.get_cell_value(ws3, 'C9')
        c9_ws4 = self.get_cell_value(ws4, 'C9')

        # 计算差值占比
        diff_ratio_c6 = ((c6_ws1 - c6_ws2) / c6_ws2) * 100 if c6_ws2 != 0 else 0
        diff_ratio_c9 = ((c9_ws3 - c9_ws4) / c9_ws4) * 100 if c9_ws4 != 0 else 0

        # 判断条件是否满足
        is_c6_condition_met = diff_ratio_c6 > 20
        is_c9_condition_met = diff_ratio_c9 > 5
        is_abnormal = is_c6_condition_met and is_c9_condition_met
        status = not is_abnormal  # 校验通过为True，异常为False

        # 添加结果
        results.append(('表1 C6 - 表2 C6 差值占比', f'{diff_ratio_c6:.2f}%', '>20%', f'{is_c6_condition_met}', True))
        results.append(('表3 C9 - 表4 C9 差值占比', f'{diff_ratio_c9:.2f}%', '>5%', f'{is_c9_condition_met}', True))
        results.append(('利率异常校验结果', '', '', f'{is_abnormal}', status))

        return results

    def check_loan_provision_coverage(self, ws1, ws2, ws3, ws4):
        """贷款拨备覆盖率校验：
        校验表1中C10的数据-表2中C10的数据大于0，
        同时表3中C6的数据除以C9的数据-表4中C6的数据除以C9的数据的百分比大于等于0，
        则校验结果为异常
        """
        results = []

        # 获取单元格值
        c10_ws1 = self.get_cell_value(ws1, 'C10')
        c10_ws2 = self.get_cell_value(ws2, 'C10')
        c6_ws3 = self.get_cell_value(ws3, 'C6')
        c9_ws3 = self.get_cell_value(ws3, 'C9')
        c6_ws4 = self.get_cell_value(ws4, 'C6')
        c9_ws4 = self.get_cell_value(ws4, 'C9')

        # 计算差值和百分比
        diff_c10 = c10_ws1 - c10_ws2

        # 避免除以零错误
        ratio_c6_c9_ws3 = (c6_ws3 / c9_ws3) * 100 if c9_ws3 != 0 else 0
        ratio_c6_c9_ws4 = (c6_ws4 / c9_ws4) * 100 if c9_ws4 != 0 else 0
        diff_ratio = ratio_c6_c9_ws3 - ratio_c6_c9_ws4

        # 判断条件是否满足
        is_c10_condition_met = diff_c10 > 0
        is_ratio_condition_met = diff_ratio >= 0
        is_abnormal = is_c10_condition_met and is_ratio_condition_met
        status = not is_abnormal  # 校验通过为True，异常为False

        # 添加结果
        results.append(('表1 C10 - 表2 C10 差值', f'{diff_c10}', '>0', f'{is_c10_condition_met}', True))
        results.append(('表3 C6/C9 - 表4 C6/C9 差值', f'{diff_ratio:.2f}%', '>=0', f'{is_ratio_condition_met}', True))
        results.append(('贷款拨备覆盖率校验结果', '', '', f'{is_abnormal}', status))

        return results

    def check_loan_items(self, ws1, ws2):
        """各项贷款核对"""
        c6_val = self.get_cell_value(ws1, 'C6')
        c8_val = self.get_cell_value(ws2, 'C8')

        return [
            ("各项贷款核对", c6_val, c8_val, c6_val == c8_val)
        ]

    def check_overdue_loans(self, ws1, ws2):
        """逾期贷款核对"""
        c14_val = self.get_cell_value(ws1, 'C14')
        sum_c24_g30 = self.sum_range(ws2, 'C24', 'G30')

        return [
            ("逾期贷款核对", c14_val, sum_c24_g30, c14_val >= sum_c24_g30)
        ]

    def check_overdue_60_days(self, ws1, ws2):
        """逾期贷款(60天以上)核对"""
        c15_val = self.get_cell_value(ws1, 'C15')
        sum_c26_g30 = self.sum_range(ws2, 'C26', 'G30')

        return [
            ("逾期贷款(60天以上)核对", c15_val, sum_c26_g30, c15_val >= sum_c26_g30)
        ]

    def check_overdue_90_days(self, ws1, ws2):
        """逾期贷款(90天以上)核对"""
        c16_val = self.get_cell_value(ws1, 'C16')
        sum_c27_g30 = self.sum_range(ws2, 'C27', 'G30')

        return [
            ("逾期贷款(90天以上)核对", c16_val, sum_c27_g30, c16_val >= sum_c27_g30)
        ]

    def check_overdue_days(self, ws1, ws2):
        """各逾期天数贷款核对"""
        results = []

        # 定义要检查的行范围
        rows = range(24, 31)  # 24到30行

        for row in rows:
            # 表1中的区域：E,F,H,I列
            sum_e_f_h_i = (
                self.get_cell_value(ws1, f'E{row}') +
                self.get_cell_value(ws1, f'F{row}') +
                self.get_cell_value(ws1, f'H{row}') +
                self.get_cell_value(ws1, f'I{row}')
            )

            # 表2中的对应P列单元格
            p_val = self.get_cell_value(ws2, f'P{row+5}')  # P29对应24行，P30对应25行...

            results.append(
                (f"逾期天数贷款核对(行{row})", sum_e_f_h_i, p_val, sum_e_f_h_i >= p_val)
            )

        return results

    def check_housing_loan_balance(self, ws1, ws2):
        """个人住房贷款期末余额核对(S67现期与基期)"""
        # 表1中c38+c39+c40的和
        sum_c38_c40 = (
            self.get_cell_value(ws1, 'C38') +
            self.get_cell_value(ws1, 'C39') +
            self.get_cell_value(ws1, 'C40')
        )

        # 表1中d38+d39+d40的和减去e36
        sum_d38_d40_minus_e36 = (
            self.get_cell_value(ws1, 'D38') +
            self.get_cell_value(ws1, 'D39') +
            self.get_cell_value(ws1, 'D40') -
            self.get_cell_value(ws1, 'E36')
        )

        # 表2中c38+c39+c40的和
        sum_ws2_c38_c40 = (
            self.get_cell_value(ws2, 'C38') +
            self.get_cell_value(ws2, 'C39') +
            self.get_cell_value(ws2, 'C40')
        )

        # 计算等式右边的值
        right_side = sum_d38_d40_minus_e36 + sum_ws2_c38_c40

        return [
            ("个人住房贷款期末余额核对", sum_c38_c40, right_side, abs(sum_c38_c40 - right_side) < 0.01)
        ]

    def check_overdue_loan_1(self, ws1, ws2):
        """逾期贷款核对1(S67和G01)"""
        # 表1中的l8+l11+l26+l36+l68的和
        sum_l8_l68 = (
            self.get_cell_value(ws1, 'L8') +
            self.get_cell_value(ws1, 'L11') +
            self.get_cell_value(ws1, 'L26') +
            self.get_cell_value(ws1, 'L36') +
            self.get_cell_value(ws1, 'L68')
        )

        # 表2中c14的值
        c14_val = self.get_cell_value(ws2, 'C14')

        return [
            ("逾期贷款核对1", sum_l8_l68, c14_val, abs(sum_l8_l68 - c14_val) < 0.01)
        ]

    def check_overdue_loan_2(self, ws1, ws2):
        """逾期贷款核对2(S67和G01)"""
        # 表1中的o8+o11+o26+o36+o68+p8+p11+p26+p36+p68的和
        sum_o_p = (
            self.get_cell_value(ws1, 'O8') +
            self.get_cell_value(ws1, 'O11') +
            self.get_cell_value(ws1, 'O26') +
            self.get_cell_value(ws1, 'O36') +
            self.get_cell_value(ws1, 'O68') +
            self.get_cell_value(ws1, 'P8') +
            self.get_cell_value(ws1, 'P11') +
            self.get_cell_value(ws1, 'P26') +
            self.get_cell_value(ws1, 'P36') +
            self.get_cell_value(ws1, 'P68')
        )

        # 表2中c15的值
        c15_val = self.get_cell_value(ws2, 'C15')

        return [
            ("逾期贷款核对2", sum_o_p, c15_val, abs(sum_o_p - c15_val) < 0.01)
        ]

    def check_overdue_loan_3(self, ws1, ws2):
        """逾期贷款核对3(S67和G01)"""
        # 表1中的p8+p11+p26+p36+p68的和
        sum_p8_p68 = (
            self.get_cell_value(ws1, 'P8') +
            self.get_cell_value(ws1, 'P11') +
            self.get_cell_value(ws1, 'P26') +
            self.get_cell_value(ws1, 'P36') +
            self.get_cell_value(ws1, 'P68')
        )

        # 表2中c16的值
        c16_val = self.get_cell_value(ws2, 'C16')

        return [
            ("逾期贷款核对3", sum_p8_p68, c16_val, abs(sum_p8_p68 - c16_val) < 0.01)
        ]

    def check_housing_loan_ratio_1(self, ws1, ws2):
        """房贷余额占比核对1(S67和G01)"""
        # 表1中c8+c11+c26+c36+c68的和
        sum_c8_c68 = (
            self.get_cell_value(ws1, 'C8') +
            self.get_cell_value(ws1, 'C11') +
            self.get_cell_value(ws1, 'C26') +
            self.get_cell_value(ws1, 'C36') +
            self.get_cell_value(ws1, 'C68')
        )

        # 表2中c7+c10的和
        sum_c7_c10 = (
            self.get_cell_value(ws2, 'C7') +
            self.get_cell_value(ws2, 'C10')
        )

        # 计算占比
        ratio = sum_c8_c68 / sum_c7_c10 if sum_c7_c10 != 0 else 0

        # 检查是否小于20%
        is_pass = ratio < 0.2

        return [
            ("房贷余额占比核对1", f"{ratio:.2%}", "20%", is_pass)
        ]

    def check_housing_loan_ratio_2(self, ws1, ws2):
        """房贷余额占比核对2(S67和G01)"""
        # 表1中c38+c39+c40的和
        sum_c38_c40 = (
            self.get_cell_value(ws1, 'C38') +
            self.get_cell_value(ws1, 'C39') +
            self.get_cell_value(ws1, 'C40')
        )

        # 表2中c7+c10的和
        sum_c7_c10 = (
            self.get_cell_value(ws2, 'C7') +
            self.get_cell_value(ws2, 'C10')
        )

        # 计算占比
        ratio = sum_c38_c40 / sum_c7_c10 if sum_c7_c10 != 0 else 0

        # 检查是否小于15%
        is_pass = ratio < 0.15

        return [
            ("房贷余额占比核对2", f"{ratio:.2%}", "15%", is_pass)
        ]

    def get_cell_value(self, ws, cell):
        """获取单元格的值，如果为空则返回0"""
        value = ws[cell].value
        return value if value is not None else 0

    def sum_range(self, ws, start_cell, end_cell):
        """计算指定范围内单元格的和"""
        start_col, start_row = self._parse_cell(start_cell)
        end_col, end_row = self._parse_cell(end_cell)

        total = 0
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                col_letter = self._col_num_to_letter(col)
                cell = f'{col_letter}{row}'
                total += self.get_cell_value(ws, cell)

        return total

    def _parse_cell(self, cell):
        """解析单元格坐标，返回(列号, 行号)"""
        import re
        match = re.match(r'([A-Z]+)(\d+)', cell)
        if not match:
            raise ValueError(f'Invalid cell format: {cell}')

        col_str, row_str = match.groups()
        row = int(row_str)

        # 计算列号 (A=1, B=2, ..., Z=26, AA=27, 等)
        col = 0
        for c in col_str:
            col = col * 26 + (ord(c.upper()) - ord('A') + 1)

        return col, row

    def _col_num_to_letter(self, num):
        """将列号转换为列字母"""
        letter = ''
        while num > 0:
            num -= 1
            letter = chr(num % 26 + ord('A')) + letter
            num = num // 26
        return letter

    def execute_single_check(self, option_id, file_path, sheet_name):
        """
        执行单表校验
        :param option_id: 校验选项ID
        :param file_path: 文件路径
        :param sheet_name: sheet名称
        :return: 校验结果列表 [(项目, 单元格值, 期望值, 是否通过), ...]
        """
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb[sheet_name]

            if option_id == 1:
                return self.check_commercial_housing_loan_1(ws)
            elif option_id == 2:
                return self.check_commercial_housing_loan_2(ws)
            elif option_id == 3:
                return self.check_loan_interest_rate(ws)
            elif option_id == 4:
                return self.check_id_numbers(ws)
            else:
                return []
        except Exception as e:
            print(f"Error during single check: {e}")
            return []

    def check_commercial_housing_loan_1(self, ws):
        """个人购买商业用房贷款核对1(S67)：校验C29与G29+H29+I29是否相等"""
        c29_val = self.get_cell_value(ws, 'C29')
        g29_val = self.get_cell_value(ws, 'G29')
        h29_val = self.get_cell_value(ws, 'H29')
        i29_val = self.get_cell_value(ws, 'I29')

        sum_g29_i29 = g29_val + h29_val + i29_val

        return [
            ("个人购买商业用房贷款核对1", c29_val, sum_g29_i29, abs(c29_val - sum_g29_i29) < 0.01)
        ]

    def check_commercial_housing_loan_2(self, ws):
        """个人购买商业用房贷款核对2(S67)：校验C29与C31+C32+C33+C34+C35是否相等"""
        c29_val = self.get_cell_value(ws, 'C29')
        c31_val = self.get_cell_value(ws, 'C31')
        c32_val = self.get_cell_value(ws, 'C32')
        c33_val = self.get_cell_value(ws, 'C33')
        c34_val = self.get_cell_value(ws, 'C34')
        c35_val = self.get_cell_value(ws, 'C35')

        sum_c31_c35 = c31_val + c32_val + c33_val + c34_val + c35_val

        return [
            ("个人购买商业用房贷款核对2", c29_val, sum_c31_c35, abs(c29_val - sum_c31_c35) < 0.01)
        ]

    def check_loan_interest_rate(self, ws):
        """基于贷款市场报价利率核对(S67)：校验D53是否等于D54+D55+D56+D57+D58"""
        d53_val = self.get_cell_value(ws, 'D53')
        d54_val = self.get_cell_value(ws, 'D54')
        d55_val = self.get_cell_value(ws, 'D55')
        d56_val = self.get_cell_value(ws, 'D56')
        d57_val = self.get_cell_value(ws, 'D57')
        d58_val = self.get_cell_value(ws, 'D58')

        sum_d54_d58 = d54_val + d55_val + d56_val + d57_val + d58_val

        return [
            ("基于贷款市场报价利率核对", d53_val, sum_d54_d58, abs(d53_val - sum_d54_d58) < 0.01)
        ]

    def check_id_numbers(self, ws):
        """证件号码核对(G1404)：校验E10到E100中每行数据是否是18位"""
        results = []

        for row in range(10, 101):
            cell = f'E{row}'
            id_number = ws[cell].value

            # 检查是否为字符串且长度为18
            is_valid = isinstance(id_number, str) and len(id_number) == 18

            results.append(
                (f"证件号码核对(行{row})", str(id_number), "18位字符串", is_valid)
            )

        return results
